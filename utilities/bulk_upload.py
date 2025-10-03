import os
import re
import sys
import yaml
import uuid
import shutil
import iiiflow
import argparse
from openpyxl import load_workbook
import asnake.logging as logging
from asnake.client import ASnakeClient
from datetime import datetime, timezone
from pathlib import PureWindowsPath, PurePosixPath

processingDir = "/backlog"
SPE_DAO = "/SPE_DAO"

# Setup ASpace connection
logging.setup_logging(filename="/logs/aspace-flask.log", filemode="a", level="INFO")
client = ASnakeClient()

PREFIXES = ["apap", "ger", "mss", "ua"]
def validate_col_id(col_id: str) -> bool:
    pattern = r'^(apap\d{3}|ger\d{3}|mss\d{3}|ua\d{3}(?:\.\d{3})?)$'
    return bool(re.match(pattern, col_id))

# Validations
ALLOWED_INPUT_FORMATS = {"jpg", "png", "ogg", "mp3", "webm"}
ALLOWED_RESOURCE_TYPES = {
    "audio", "bound volume", "dataset", "document", "image", "map",
    "mixed materials", "pamphlet", "periodical", "slides", "video", "other"
}
ALLOWED_LICENSES = {"CC BY", "CC BY-NC-SA", "Unknown"}
ALLOWED_BEHAVIORS = {"paged", "individuals", "continuous"}

# License and rights lookup
LICENSE_LOOKUP = {
    "CC BY": "https://creativecommons.org/licenses/by/4.0/",
    "CC BY-NC-SA": "https://creativecommons.org/licenses/by-nc-sa/4.0/",
}
RIGHTS_FALLBACK = "https://rightsstatements.org/page/InC-EDU/1.0/"

parser = argparse.ArgumentParser()
parser.add_argument("package", help="ID for package you are processing, e.g. 'ua950.012_Xf5xzeim7n4yE6tjKKHqLM'")
parser.add_argument("-s", "--sheet", required=True,
                    help="Path (relative to the package's metadata folder) of the bulk upload spreadsheet. "
                         "Example: sheet.xlsx or sub/folder/sheet.xlsx")

args = parser.parse_args()

if not os.path.isdir(SPE_DAO):
    raise FileNotFoundError("Error: Missing /SPE_DAO mount.")

# Extract base ID from package name
if "_" in args.package:
    ID = args.package.split("_")[0]
elif "-" in args.package:
    ID = args.package.split("-")[0]
else:
    raise Exception(f"Error: {args.package} is not a valid processing package.")

if not validate_col_id(ID):
    raise ValueError(f"Invalid collection ID: {ID}")

# Base metadata path
metadata_path = os.path.join(processingDir, ID, args.package, "metadata")
if not os.path.isdir(metadata_path):
    raise FileNotFoundError(f"Error: missing path {metadata_path}")
derivatives_path = os.path.join(processingDir, ID, args.package, "derivatives")
masters_path = os.path.join(processingDir, ID, args.package, "masters")

# Normalize sheet path (handle Windows-style paths too)
sheet_arg = args.sheet
if "\\" in sheet_arg:
    win_path = PureWindowsPath(sheet_arg)
    sheet_rel = PurePosixPath(*win_path.parts)
else:
    sheet_rel = os.path.normpath(sheet_arg)

sheet_path = os.path.join(metadata_path, sheet_rel)

if not os.path.isfile(sheet_path):
    raise FileNotFoundError(f"Error: sheet file not found at {sheet_path}")

print(f"Reading bulk upload sheet {sheet_path}...")

wb = load_workbook(sheet_path, data_only=True)
ws = wb.active

# Get headers from row 6
headers = [str(cell.value).strip() if cell.value else "" for cell in ws[6]]

required_cols = [
    "ArchivesSpace ID", "Title", "File Paths",
    "Input Format", "Resource Type", "Behavior",
    "License/Rights"
]

for col in required_cols:
    if col not in headers:
        raise ValueError(f"Missing required column: {col}")

# Build a map of all headers
col_index = {h: i for i, h in enumerate(headers) if h}

errors = []
records = []

for row_num, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
    if all(v is None for v in row):
        continue

    title = str(row[col_index["Title"]]).strip() if row[col_index["Title"]] else ""
    aspace_id = str(row[col_index["ArchivesSpace ID"]]).strip() if row[col_index["ArchivesSpace ID"]] else ""
    original_file = str(row[col_index["Original file"]]).strip() if row[col_index["Original file"]] else ""

    if not (title and aspace_id):
        continue  # skip incomplete rows

    # Ensure these fileds have text
    missing_fields = []
    for field in ["File Paths", "Input Format", "Resource Type", "Behavior", "License/Rights"]:
        value = str(row[col_index[field]]).strip() if row[col_index[field]] else ""
        if not value:
            if field == "Behavior":
                fmt = str(row[col_index["Input Format"]]).strip() if row[col_index["Input Format"]] else ""
                if "png" in fmt.lower() or "jpg" in fmt.lower():
                    missing_fields.append(field)
            else:
                missing_fields.append(field)
    if missing_fields:
        errors.append(f"Missing required fields {missing_fields} in row {row_num}")

    # Handle File Paths
    file_path_raw = row[col_index["File Paths"]]
    file_path = ""
    if file_path_raw:
        file_path = str(file_path_raw).strip()
        if "|" in file_path:
            errors.append(f"File path does not support multiple paths or pipes (|), (row {row_num})")
        full_path = os.path.normpath(os.path.join(derivatives_path, file_path))
        if not os.path.isdir(full_path):
            full_path = os.path.normpath(os.path.join(masters_path, file_path))
            if not os.path.isdir(full_path):
                errors.append(f"Missing path: {full_path} (row {row_num})")

    # Check original file path
    if len(original_file) > 0:
        derivative_original = os.path.join(derivatives_path, os.path.normpath(original_file))
        if not os.path.isfile(derivative_original):
            masters_original = os.path.join(masters_path, os.path.normpath(original_file))
            if not os.path.isfile(masters_original):
                errors.append(f"Original file {original_file} does not exist in {derivative_original} or {masters_original} (row {row_num}).")

    # Check if SPE_DAO path already exists
    object_path = os.path.join(SPE_DAO, ID, aspace_id)
    if os.path.isdir(aspace_id):
        errors.append(f"SPE_DAO object path '{object_path}' already exists for row {row_num}")

    # Validate controlled vocab fields
    input_fmt = str(row[col_index["Input Format"]]).strip().lower() if row[col_index["Input Format"]] else ""
    if input_fmt and input_fmt.lower() not in ALLOWED_INPUT_FORMATS:
        errors.append(f"Invalid Input Format '{input_fmt}' in row {row_num}")

    res_type = str(row[col_index["Resource Type"]]).strip() if row[col_index["Resource Type"]] else ""
    if res_type and res_type.lower() not in ALLOWED_RESOURCE_TYPES:
        errors.append(f"Invalid Resource Type '{res_type}' in row {row_num}")

    license_val = str(row[col_index["License/Rights"]]).strip() if row[col_index["License/Rights"]] else ""
    if license_val and license_val not in ALLOWED_LICENSES:
        errors.append(f"Invalid License '{license_val}' in row {row_num}")

    behavior = str(row[col_index["Behavior"]]).strip().lower() if row[col_index["Behavior"]] else ""
    if behavior and behavior.lower() not in ALLOWED_BEHAVIORS:
        errors.append(f"Invalid Behavior '{behavior}' in row {row_num}")

    records.append({
        "ArchivesSpace ID": aspace_id,
        "Title": title,
        "File Path": full_path,
        "Original file": original_file,
        "Input Format": input_fmt,
        "Resource Type": res_type,
        "License/Rights": license_val,
        "Behavior": behavior,
    })

if errors:
    print("Validation errors found:")
    for e in errors:
        print(" -", e)
    print ("Exiting.")
    sys.exit()
else:
    print("All records validated successfully!")

print(f"Parsed {len(records)} valid records.")

# Loop though validated sheet to upload files
error_count = 0
for rec in records:
    error_switch = False
    title = rec["Title"]
    aspace_id = rec["ArchivesSpace ID"]
    file_path = rec["File Path"]
    original_file = rec["Original file"]
    res_type = rec["Resource Type"]
    license = rec["License/Rights"]
    input_fmt = rec["Input Format"]
    behavior = rec["Behavior"]

    print(f"\tProcessing {title} ({aspace_id})...")

    collectionDir = os.path.join (SPE_DAO, ID)
    if not os.path.isdir(collectionDir):
        os.mkdir(collectionDir)

    object_path = os.path.join(collectionDir, aspace_id)
    if os.path.isdir(aspace_id):
        raise FileExistsError(f"Error: access path {object_path} already present.")
    os.mkdir(object_path)

    # Build metadata.yml
    metadata = {
        "preservation_package": args.package,
        "resource_type": res_type.title(),
        "date_uploaded": datetime.now(timezone.utc).isoformat(),
    }
    if len(original_file) > 0:
        metadata["original_file"] = original_file
    if len(behavior) > 0:
        metadata["behavior"] = behavior.lower()

    lic = license_val.strip() if license_val else ""
    if lic in LICENSE_LOOKUP:
        metadata["license"] = LICENSE_LOOKUP[lic]
    else:
        # Treat "Unknown" or empty as rights statement fallback
        metadata["rights_statement"] = RIGHTS_FALLBACK

    metadata_path = os.path.join(object_path, "metadata.yml")
    with open(metadata_path, "w", encoding="utf-8") as f:
        yaml.dump(metadata, f, sort_keys=False, allow_unicode=True)

    print(f"\tCreated metadata.yml at {metadata_path}")

    format_path = os.path.join(object_path, input_fmt.lower())
    os.mkdir(format_path)
    for file in os.scandir(file_path):
        #print (f"\t\tMoving {file.path} to {format_path}...")
        print (f"\t\tMoving {file.path}...")
        shutil.copy2(file.path, format_path)

    # make thumbnail
    #print ("\tCreating thumbnail")
    iiiflow.make_thumbnail(ID, aspace_id)

    img_formats = ["png", "jpg"]
    if ".pdf" in original_file.lower():
        print (f"\tMoving {original_file} to package as original file...")
        original_path = os.path.join(derivatives_path, os.path.normpath(original_file))
        if not os.path.isfile(original_path):
            original_path = os.path.join(masters_path, os.path.normpath(original_file))
            if not os.path.isfile(original_path):
                raise FileNotFoundError(f"Missing original file {original_file}.")
        pdf_path = os.path.join(object_path, "pdf")
        os.mkdir(pdf_path)
        shutil.copy2(original_path, pdf_path)
    elif input_fmt.lower() in img_formats:
        print ("\tCreating alternative PDF...")
        iiiflow.create_pdf(ID, aspace_id)

    # Create pyramidal tifs
    if input_fmt.lower() in img_formats:
        print ("\tCreating pyramidal tifs (.ptifs)...")
        iiiflow.create_ptif(ID, aspace_id)

    # OCR/transcription
    if input_fmt.lower() in img_formats:
        print ("\tRecognizing text...")
        iiiflow.create_hocr(ID, aspace_id)
    elif input_fmt.lower() in ("ogg", "mp3", "webm"):
        print ("\tTranscribing...")
        iiiflow.create_transcription(ID, aspace_id)

    # Index HOCR
    if input_fmt.lower() in img_formats:
        print ("\tIndexing text for content search...")
        iiiflow.index_hocr_to_solr(ID, aspace_id)

    # Create manifest
    print ("\tGenerating IIIF manifest...")
    iiiflow.create_manifest(ID, aspace_id)

    print ("\tAdding digital object record to ArchivesSpace...")

    # get url root from iiiflow config
    with open("/root/.iiiflow.yml", "r") as config_file:
        config = yaml.safe_load(config_file)
    manifest_url_root = config.get("manifest_url_root")
    dao_url = f"{manifest_url_root}/{ID}/{aspace_id}/manifest.json"

    file_version = {
        "jsonmodel_type": "file_version",
        "publish": True,
        "is_representative": True,
        "file_uri": dao_url,
        "use_statement": "",
        "xlink_actuate_attribute": "none",
        "xlink_show_attribute": "embed",
    }
    dao_uuid = str(uuid.uuid4())
    dao_title = "Online object uploaded typically on user request."

    dao_object = {
        "jsonmodel_type": "digital_object",
        "publish": True,
        "is_representative": True,
        "external_ids": [],
        "subjects": [],
        "linked_events": [],
        "extents": [],
        "dates": [],
        "external_documents": [],
        "rights_statements": [],
        "linked_agents": [],
        "file_versions": [file_version],
        "restrictions": False,
        "notes": [],
        "linked_instances": [],
        "title": dao_title,
        "language": "",
        "digital_object_id": dao_uuid,
    }

    # Upload new digital object
    ref = client.get("repositories/2/find_by_id/archival_objects?ref_id[]=" + aspace_id).json()
    item = client.get(ref["archival_objects"][0]["ref"]).json()
    new_dao = client.post("repositories/2/digital_objects", json=dao_object)
    if new_dao.status_code != 200:
        error_switch = True
        print (json.dumps(dao_object, indent=4))
        print (f"Failed adding digital object record --> {new_dao.status_code}")
        print (new_dao.reason)
    else:
        dao_uri = new_dao.json()["uri"]
        print (f"Added digital object record {dao_uri} --> {new_dao.status_code}")

    # Attach new digital object instance to archival object
    dao_link = {
        "jsonmodel_type": "instance",
        "digital_object": {"ref": dao_uri},
        "instance_type": "digital_object",
        "is_representative": True,
    }

    # Ensure all other instances are is_representative = False
    for instance in item["instances"]:
        instance["is_representative"] = False
    item["instances"].append(dao_link)

    """
    if args.processing and len(args.processing) > 0:
        processing_note = {
            "jsonmodel_type": "note_multipart",
            "type": "processinfo",
            "subnotes": [
                {
                    "jsonmodel_type": "note_text",
                    "content": args.processing,
                    "publish": True
                }
            ],
            "publish": True
        }
        item["notes"].append(processing_note)
        print (f"Added processing note.")
    """
    update_item = client.post(item["uri"], json=item)
    if update_item.status_code == 200:
        print (f"Updated archival object record --> {update_item.status_code}")
    else:
        error_switch = True
        print (json.dumps(item, indent=4))
        print (f"Failed updating archival object record --> {update_item.status_code}")
        print (update_item.reason)

    #print ("Indexing record in ArcLight... (skipping)")
    
    if error_switch == False:
        print ("Success!")
    else:
        error_count += 1
        print ("Uploaded with errors.")
    print (f"Check out digital object at:")
    print (f"https://media.archives.albany.edu?manifest={dao_url}")


if error_count == len(records):
    print (f"Success! {len(records)} records upload successfully.")
else:
    print (f"Errors: {error_count} out of {records} uploaded successfully.")
