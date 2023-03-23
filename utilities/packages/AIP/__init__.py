import os
import csv
import time
import bagit
import shutil
from datetime import datetime
from subprocess import Popen, PIPE

class ArchivalInformationPackage:

    def __init__(self):
        self.excludeList = ["thumbs.db", "desktop.ini", ".ds_store"]
        self.aipPath= "/Archives/AIP"
        
    def load(self, path):
        if not os.path.isdir(path):
            raise Exception("ERROR: " + str(path) + " is not a valid AIP. You may want to create a AIP with .create().")

        self.bag = bagit.Bag(path)
        self.accession = os.path.basename(path)
        if "_" in self.accession:
            self.colID = self.accession.split("_")[0]
        else:
            self.colID = self.accession.split("-")[0]
        self.bagDir = (path)
        self.data = os.path.join(self.bagDir, "data")        
    
    def create(self, colID, accession):
        
        metadata = {\
        'Bag-Type': 'AIP', \
        'Bagging-Date': str(datetime.now().isoformat()), \
        'Posix-Date': str(time.time()), \
        'BagIt-Profile-Identifier': 'https://archives.albany.edu/static/bagitprofiles/aip-profile-v0.1.json', \
        }
        
        self.accession = accession
        self.colID = accession.split("_")[0]
        if not os.path.isdir(os.path.join(self.aipPath, colID)):
            os.mkdir(os.path.join(self.aipPath, colID))
        self.bagDir = os.path.join(self.aipPath, colID, accession)
        os.mkdir(self.bagDir)
        metadata["Bag-Identifier"] = accession
        metadata["Collection-Identifier"] = colID
        
        self.bag = bagit.make_bag(self.bagDir, metadata)
        self.data = os.path.join(self.bagDir, "data")
        
    def clean(self):
        for root, dirs, files in os.walk(self.data):
            for file in files:
                if file.lower() in self.excludeList:
                    filePath = os.path.join(root, file)
                    print ("removing " + filePath)
                    os.remove(filePath)
    
    def addMetadata(self, hyraxData):
        headers = ["Type", "URIs", "File Paths", "Accession", "Collecting Area", "Collection Number", "Collection", \
        "ArchivesSpace ID", "Record Parents", "Title", "Description", "Date Created", "Resource Type", "License", \
        "Rights Statement", "Subjects", "Whole/Part", "Processing Activity", "Extent", "Language"]
        metadataPath = os.path.join(self.bagDir, "metadata")
        if not os.path.isdir(metadataPath):
            os.mkdir(metadataPath)
        metadataFile = os.path.join(metadataPath, self.accession + ".tsv")
        addHeaders = False
        if not os.path.isfile(metadataFile):
            addHeaders = True
        outfile = open(metadataFile, "a")
        writer = csv.writer(outfile, delimiter='\t', lineterminator='\n')
        if addHeaders == True:
            writer.writerow(headers)
        writer.writerow(hyraxData)
        outfile.close()
        
    def addFile(file):
        dataPath = os.path.join(self.bagDir, "data")
        if not os.path.isdir(metadataPath):
            os.mkdir(dataPath)
        shutil.copy2(file, dataPath)
        
    def copyRsync(self, source, destination, retry=0):
        retry += 1
        if retry < 6:
            cmd = ["rsync", "-arv", "--partial", source, destination]
            print ("Copy attempt " + str(retry) + " at " + str(datetime.now()))
            print ("Running " + " ".join(cmd))
            p = Popen(cmd, stdout=PIPE, stderr=PIPE)
            stdout, stderr = p.communicate()
            if p.returncode != 0:
                print (stdout)
                print (stderr)
                print ("Copy failed at " + str(datetime.now()))
                print ("Retrying...")
                self.copyRsync(source, destination, retry)
            else:
                print ("Success!")
                print("Copy completed at " + str(datetime.now()))
                print (stdout)
                if len(stderr) > 0:
                    print (stderr)
        else:
            print ("Failed to copy in 5 attempts")
            raise ValueError("Failed to copy in 5 attempts")
    
    def packageFiles(self, type, dir):
        allowed = ["derivatives", "masters"]
        if not type in allowed:
            raise Exception("ERROR: " + str(type) + " is not a valid subfolder in the AIP /data directory.")
        if not os.path.isdir(dir):
            raise Exception("ERROR: " + str(dir) + " is not a valid path.")
        else:
            dest = os.path.join(self.data, type)
            if not os.path.isdir(dest):
                os.mkdir(dest)
            print(datetime.now())
            # Move files and folders to AIP
            self.copyRsync(os.path.join(dir, ""), os.path.join(dest, ""))
                    
    def packageMetadata(self, dir, subfolder=None):
        if isinstance(dir, (list,)):
            pathList = dir 
        elif not os.path.isdir(dir):
            raise Exception("ERROR: " + str(dir) + " is not a valid path.")
        else:
            pathList = []
            for item in os.listdir(dir):
                pathList.append(os.path.join(dir, item))
        
        metadataDir = os.path.join(self.bag.path, "metadata")
        if not os.path.isdir(metadataDir):
            os.mkdir(metadataDir)
        if subfolder is None:
            dest = metadataDir
        else:
            dest = os.path.join(metadataDir, subfolder)
            if not os.path.isdir(dest):
                os.mkdir(dest)
            
        # Move files and folders to metadataDir
        for thing in pathList:
            print ("Moving " + thing)
            if os.path.isfile(thing):
                if not thing.lower() in self.excludeList:
                    shutil.copy2(thing, dest)
            elif os.path.isdir(thing):
                newDir = os.path.join(dest, os.path.basename(thing))
                shutil.copytree(thing, newDir)
                #pass
            else:
                raise Exception("ERROR: " + str(thing) + " is not a valid path.")
        
    def addSIPData(self, sipDir):
        if not os.path.isdir(sipDir):
            raise Exception("ERROR: SIP " + str(sipDir) + " is not a valid path.")
        dest = os.path.join(self.bag.path, "SIP")
        if not os.path.isdir(dest):
            os.mkdir(dest)
        for file in os.listdir(sipDir):
            filePath = os.path.join(sipDir, file)
            if os.path.isfile(filePath):
                shutil.copy2(filePath, dest)
    
    def checkSIPManifest(self, algorithm="sha256"):
        import hashlib
        BUF_SIZE = 512 * 1024
        
        sipManifestPath = os.path.join(self.bagDir, "SIP", "manifest-" + algorithm.lower().strip() + ".txt")
        if not os.path.isfile(sipManifestPath):
            print ("Cannot validate against SIP manifest file, no sha256 SIP manifest file is present")
            return False
        else:
            actualFiles = []
            mastersDir = os.path.join(self.data, "masters")
            for root, dirs, files in os.walk(mastersDir):
                for file in files:
                
                    relPath = "data" + os.path.join(root, file).split("data")[1]
                    #print (relPath)
                    actualFiles.append(relPath)
            success = True
            noFileErrors = []
            badFileErrors = []
            sipManifestFile = open(sipManifestPath, "r")
            for line in sipManifestFile.read().splitlines():
                hash, oldFilePath = line.split("  ")
                filePath = os.path.join("data", "masters", oldFilePath.split("data/")[1])
                if not os.path.isfile(os.path.join(self.bagDir, filePath)):
                    noFileErrors.append(filePath)
                else:
                    actualFiles.remove(filePath)
                    if algorithm.lower() == "md5":
                        fileHash = hashlib.md5()
                    elif algorithm.lower() == "sha256":
                        fileHash = hashlib.sha256()
                    elif algorithm.lower() == "sha512":
                        fileHash = hashlib.sha512()
                    else:
                        raise ValueError("Error: Hash algorithm " + algorithm + " not supported.")
                    with open(os.path.join(self.bagDir, filePath), 'rb') as f:
                        while True:
                            data = f.read(BUF_SIZE)
                            if not data:
                                break
                            fileHash.update(data)
                    if not str(fileHash.hexdigest()) == hash:
                        badFileErrors.append(filePath)
                    
            if len(noFileErrors) > 0:
                success = False
                print ("Missing files defined in SIP manifest:")
                print ("\t" + "\n\t".join(noFileErrors))
            if len(badFileErrors) > 0:
                success = False
                print ("Files defined in SIP manifest have changed:")
                print ("\t" + "\n\t".join(badFileErrors))
            if len(actualFiles) > 0:
                success = False
                print ("Additional files not in SIP manifest:")
                print ("\t" + "\n\t".join(actualFiles))
            sipManifestFile.close()
            
            return success
                
        
    
    def size(self):
        suffixes = ['B', 'KB', 'MB', 'GB', 'TB', 'PB']
        bytes, fileCount = self.bag.info["Payload-Oxum"].split(".")
        dirSize = int(bytes)
        i = 0
        while dirSize >= 1024 and i < len(suffixes)-1:
            dirSize /= 1024.
            i += 1
        f = ('%.2f' % dirSize).rstrip('0').rstrip('.')
        return [f, suffixes[i], fileCount]
        
    def extentLog(self, logFile):
        import openpyxl
        if os.path.isfile(logFile):
            wb = openpyxl.load_workbook(filename=logFile, read_only=False)
            sheet = wb.active
            startRow = int(sheet.max_row) + 1
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet["A1"] = "Date"
            sheet["B1"] = "Collection ID"
            sheet["C1"] = "Type"
            sheet["D1"] = "Package"
            sheet["E1"] = "Files"
            sheet["F1"] = "Extent"
            sheet["G1"] = "Extent (MB)"
            startRow = 2
            
        packageSize = self.size()
        sheet["A" + str(startRow)] = self.bag.info["Bagging-Date"]
        sheet["B" + str(startRow)] = self.bag.info["Collection-Identifier"]
        sheet["C" + str(startRow)] = self.bag.info["Bag-Type"]
        sheet["D" + str(startRow)] = self.bag.info["Bag-Identifier"]
        sheet["E" + str(startRow)] = int(packageSize[2])
        sheet["F" + str(startRow)] = str(packageSize[0]) + " " + str(packageSize[1])
        sheet["G" + str(startRow)] = int(int(self.bag.info["Payload-Oxum"].split(".")[0]) / 1048576)
            
        wb.save(filename=logFile)